import glob
import os

import openpyxl
import csv

TARGET_FILE = glob.glob("�ձ���/*.xlsx")[-1]


def read_xls_file(file_path=TARGET_FILE):
    """��ȡxlsx�ļ�����ӡ
    :param file_path: Excel�ļ�·��
    """
    # ��Excel�ļ�
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # ��ӡ����������
    for sheet_name in wb.sheetnames:
        print(sheet_name)
        # ��ӡ��Ԫ������
        sheet = wb[sheet_name]
        for row in sheet.rows:
            for cell in row:
                print(cell.value, end="\t")
            print()


def get_target_data_list(file_path=TARGET_FILE):
    """
    ��ȡxlsx�ļ���������ڡ���н������ÿ����ҵ����ÿ���ܳɱ���\n
    :param file_path: Excel�ļ�·��
    :return: date_list: �����б�; work_people_num_list: ��н�����ڣ������б�; daily_performance_list: ÿ��ҵ���б�; daily_cost_list: ÿ�ճɱ��б�; daily_profit_list: ÿ�������б�
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)  # ��Excel�ļ�,data_only=True����ȡֵ����ȡ��ʽ
    date_list = []  # 1. ���ڣ�Ҳ�ǹ���������
    work_people_num_list = []  # 2. ������������н����
    daily_performance_list = []  # 3. ÿ����ҵ���б�
    daily_cost_list = []  # 4. ÿ�ճɱ��б�
    daily_profit_list = []  # 5. ÿ�������б�
    for sheet_name in wb.sheetnames:
        # 1. ��ȡ����
        date_ = sheet_name  # ������������������
        date_list.append(sheet_name)
        print(f"���ڣ�{sheet_name}")  # ��ӡ���������ƣ����ڣ�

        sheet = wb[sheet_name]  # ��ȡ������
        # 2. ��ȡÿ�յ�н����
        one_sheet_work_people_num = sheet['R5'].value  # һ�ű�һ�죩�ĵ�н����
        print(f"��н������{one_sheet_work_people_num}��")
        work_people_num_list.append(one_sheet_work_people_num)
        # 3. ��ȡÿ����ҵ��
        one_sheet_daily_performance = sheet['R6'].value  # һ�ű�һ�죩����ҵ��
        daily_performance_list.append(one_sheet_daily_performance)
        print(f"ÿ����ҵ����{one_sheet_daily_performance}")
        # 4. ��ȡÿ���ܳɱ�
        one_sheet_daily_cost = sheet['R7'].value
        daily_cost_list.append(one_sheet_daily_cost)
        print(f"ÿ���ܳɱ���{one_sheet_daily_cost}")

        # 5. ��ȡÿ��������
        one_sheet_daily_profit = sheet['R8'].value
        daily_profit_list.append(one_sheet_daily_profit)
        print(f"ÿ������{one_sheet_daily_profit}")

        print()
    return date_list, work_people_num_list, daily_performance_list, daily_cost_list, daily_profit_list


def csv_to_excel(csv_filename, excel_filename):
    """
    ��csv�ļ�ת��Ϊexcel�ļ���\n
    :param csv_filename: ��ת����csv�ļ������ơ�
    :param excel_filename: Ŀ�����ɵ�Excel�ļ������ơ�
    :return:
    """
    # 1. ��ȡcsv�ļ�
    csv_data = []
    with open(csv_filename) as f:
        csv_data = [row for row in csv.reader(f)]
    # 2. д��Excel�ļ���
    workbook = openpyxl.workbook.Workbook()
    worksheet = workbook.active
    for row in csv_data:
        worksheet.append(row)
    workbook.save(excel_filename)


def main():
    """���ɺ�������Ϊ�������������������н顣��Ҫ�������ͨ�������������߼���
    """
    # 1. ��ȡ���ڡ���н���������ڵ���������ÿ����ҵ����ÿ���ܳɱ���ÿ������
    date_list, work_people_num_list, daily_performance_list, daily_cost_list, daily_profit_list = get_target_data_list()
    # 2. �����ݰ�ָ����ʽ������csv�ļ���
    date_list_ = [""] + date_list
    work_people_num_list_ = ["��н����"] + work_people_num_list
    daily_performance_list_ = ["��ҵ��"] + daily_performance_list
    daily_cost_list_ = ["�ܳɱ�"] + daily_cost_list
    daily_profit_list_ = ["����"] + daily_profit_list
    with open("ͳ�ƻ���.csv", "w", newline="", encoding="gbk") as f:
        writer = csv.writer(f)
        writer.writerow(date_list_)
        writer.writerow(work_people_num_list_)
        writer.writerow(daily_performance_list_)
        writer.writerow(daily_cost_list_)
        writer.writerow(daily_profit_list_)
    # 3. ��csv�ļ�ת����excel�ļ�
    csv_filename = "ͳ�ƻ���.csv"
    excel_filename = f"{date_list[0]}��{date_list[-1]}����ͳ�ƻ���.xlsx"
    csv_to_excel(csv_filename, excel_filename)
    # 4. ɾ��csv�ļ�
    os.remove("ͳ�ƻ���.csv")


if __name__ == "__main__":
    main()
# ���ο�������£�
# https://zhuanlan.zhihu.com/p/342422919
# https://juejin.cn/post/7109795200611909669

